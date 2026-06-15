import io
from datetime import date, datetime, timezone, timedelta
from typing import Optional
import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, status
from sqlalchemy import func as sqlfunc
from sqlalchemy.orm import Session
from .. import models, schemas
from ..database import get_db
from ..services.auth_service import require_authenticated

router = APIRouter(prefix="/positions", tags=["positions"], dependencies=[Depends(require_authenticated)])

STATUS_VALUES = {"planned", "in_progress", "completed", "delayed", "cancelled"}


def _parse_date(val) -> Optional[date]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _get_position(db: Session, position_id: int) -> models.SchedulePosition:
    pos = db.query(models.SchedulePosition).filter(models.SchedulePosition.id == position_id).first()
    if not pos:
        raise HTTPException(status_code=404, detail="Position nicht gefunden")
    return pos


@router.get("/version/{version_id}", response_model=list[schemas.PositionResponse])
def list_positions(version_id: int, db: Session = Depends(get_db)):
    version = db.query(models.ScheduleVersion).filter(models.ScheduleVersion.id == version_id).first()
    if not version:
        raise HTTPException(status_code=404, detail="Version nicht gefunden")
    return db.query(models.SchedulePosition).filter(
        models.SchedulePosition.version_id == version_id
    ).order_by(models.SchedulePosition.sort_order).all()


@router.post("/", response_model=schemas.PositionResponse, status_code=status.HTTP_201_CREATED)
def create_position(data: schemas.PositionCreate, db: Session = Depends(get_db)):
    version = db.query(models.ScheduleVersion).filter(models.ScheduleVersion.id == data.version_id).first()
    if not version:
        raise HTTPException(status_code=404, detail="Version nicht gefunden")
    pos = models.SchedulePosition(**data.model_dump())
    db.add(pos)
    db.commit()
    db.refresh(pos)
    return pos


@router.put("/{position_id}", response_model=schemas.PositionResponse)
def update_position(position_id: int, data: schemas.PositionUpdate, db: Session = Depends(get_db)):
    pos = db.query(models.SchedulePosition).filter(models.SchedulePosition.id == position_id).first()
    if not pos:
        raise HTTPException(status_code=404, detail="Position nicht gefunden")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(pos, key, value)
    db.commit()
    db.refresh(pos)
    return pos


@router.delete("/{position_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_position(position_id: int, db: Session = Depends(get_db)):
    pos = db.query(models.SchedulePosition).filter(models.SchedulePosition.id == position_id).first()
    if not pos:
        raise HTTPException(status_code=404, detail="Position nicht gefunden")
    db.delete(pos)
    db.commit()


@router.post("/version/{version_id}/reorder", status_code=status.HTTP_204_NO_CONTENT)
def reorder_positions(version_id: int, order: list[int], db: Session = Depends(get_db)):
    for idx, pos_id in enumerate(order):
        db.query(models.SchedulePosition).filter(
            models.SchedulePosition.id == pos_id,
            models.SchedulePosition.version_id == version_id,
        ).update({"sort_order": idx})
    db.commit()


@router.post("/version/{version_id}/import", response_model=schemas.ImportResult)
async def import_positions(
    version_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    version = db.query(models.ScheduleVersion).filter(models.ScheduleVersion.id == version_id).first()
    if not version:
        raise HTTPException(status_code=404, detail="Version nicht gefunden")

    content = await file.read()
    filename = (file.filename or "").lower()
    errors: list[str] = []
    imported = 0
    skipped = 0

    try:
        if filename.endswith((".xlsx", ".xls")):
            df = pd.read_excel(io.BytesIO(content), dtype=str)
        elif filename.endswith(".csv"):
            for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
                try:
                    df = pd.read_csv(io.BytesIO(content), dtype=str, encoding=enc, sep=None, engine="python")
                    break
                except Exception:
                    continue
            else:
                raise ValueError("CSV konnte nicht gelesen werden")
        else:
            raise ValueError("Nur .xlsx, .xls und .csv werden unterstützt")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    df.columns = [str(c).strip().lower().replace(" ", "_").replace("-", "_") for c in df.columns]

    col_map = {
        "pos_number": ["pos", "pos.", "pos_nr", "pos._nr.", "pos_number", "position", "nummer", "nr", "nr."],
        "title": ["title", "bezeichnung", "beschreibung", "name", "vorgang", "aufgabe", "task"],
        "start_date": ["start", "start_date", "startdatum", "beginn", "von", "anfang"],
        "end_date": ["end", "end_date", "enddatum", "ende", "bis", "fertig"],
        "duration_days": ["duration", "dauer", "duration_days", "dauer_tage", "tage", "dauer_(tage)"],
        "responsible": ["responsible", "verantwortlich", "zuständig", "leiter", "person"],
        "trade": ["trade", "gewerk", "fachbereich", "bereich"],
        "status": ["status"],
        "progress": ["progress", "fortschritt", "fertigstellung", "%", "fortschritt_(%)"],
    }

    def find_col(target: str) -> Optional[str]:
        candidates = col_map.get(target, [target])
        for c in candidates:
            if c in df.columns:
                return c
        return None

    max_order = db.query(models.SchedulePosition).filter(
        models.SchedulePosition.version_id == version_id
    ).count()

    for idx, row in df.iterrows():
        row_num = int(idx) + 2
        try:
            title_col = find_col("title")
            if not title_col or pd.isna(row.get(title_col, None)) or str(row.get(title_col, "")).strip() == "":
                skipped += 1
                continue

            def get(col_name: str) -> Optional[str]:
                c = find_col(col_name)
                if c is None:
                    return None
                v = row.get(c)
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    return None
                return str(v).strip() or None

            raw_status = get("status")
            status_val = raw_status.lower().replace(" ", "_") if raw_status else "planned"
            if status_val not in STATUS_VALUES:
                status_val = "planned"

            raw_progress = get("progress")
            progress_val = 0.0
            if raw_progress:
                try:
                    progress_val = float(raw_progress.replace("%", "").replace(",", "."))
                    if progress_val > 1:
                        progress_val = progress_val / 100
                except ValueError:
                    progress_val = 0.0

            raw_dur = get("duration_days")
            dur_val = None
            if raw_dur:
                try:
                    dur_val = int(float(raw_dur.replace(",", ".")))
                except ValueError:
                    dur_val = None

            pos = models.SchedulePosition(
                version_id=version_id,
                pos_number=get("pos_number"),
                title=str(row[title_col]).strip(),
                start_date=_parse_date(row.get(find_col("start_date") or "")),
                end_date=_parse_date(row.get(find_col("end_date") or "")),
                duration_days=dur_val,
                responsible=get("responsible"),
                trade=get("trade"),
                status=status_val,
                progress=progress_val,
                sort_order=max_order + idx,
            )
            db.add(pos)
            imported += 1
        except Exception as e:
            errors.append(f"Zeile {row_num}: {e}")

    db.commit()
    return schemas.ImportResult(imported=imported, skipped=skipped, errors=errors)


@router.get("/version/{version_id}/export")
def export_positions(version_id: int, db: Session = Depends(get_db)):
    from fastapi.responses import StreamingResponse
    version = db.query(models.ScheduleVersion).filter(models.ScheduleVersion.id == version_id).first()
    if not version:
        raise HTTPException(status_code=404, detail="Version nicht gefunden")

    positions = db.query(models.SchedulePosition).filter(
        models.SchedulePosition.version_id == version_id
    ).order_by(models.SchedulePosition.sort_order).all()

    pos_number_by_id = {p.id: p.pos_number for p in positions}

    rows = []
    for p in positions:
        parent_pos_nr = pos_number_by_id.get(p.parent_id) if p.parent_id else None
        rows.append({
            "Pos.-Nr.": p.pos_number,
            "Bezeichnung": p.title,
            "Typ": p.typ,
            "Beginn": p.start_date.strftime("%d.%m.%Y") if p.start_date else "",
            "Ende": p.end_date.strftime("%d.%m.%Y") if p.end_date else "",
            "Dauer (Tage)": p.duration_days,
            "Verantwortlich": p.responsible,
            "Gewerk": p.trade,
            "Status": p.status,
            "Fortschritt (%)": int(p.progress * 100) if p.progress else 0,
            "Übergeordnete Pos.-Nr.": parent_pos_nr or "",
        })

    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Terminplan")
    buf.seek(0)

    safe_name = "".join(c for c in version.name if c.isalnum() or c in " -_").strip()
    filename = f"Terminplan_V{version.version_number}_{safe_name}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/template")
def download_template():
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Terminplan"

    headers = [
        "Pos.-Nr.", "Bezeichnung", "Typ", "Beginn", "Ende",
        "Dauer (Tage)", "Verantwortlich", "Gewerk", "Status",
        "Fortschritt (%)", "Übergeordnete Pos.-Nr.",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    examples = [
        ["1", "Gesamtprojekt", "sammelvorgang", "01.01.2025", "31.12.2025", "", "Muster GmbH", "", "planned", 0, ""],
        ["1.1", "Phase 1: Vorbereitung", "sammelvorgang", "01.01.2025", "31.03.2025", "", "Muster GmbH", "Planung", "planned", 0, "1"],
        ["1.1.1", "Planung abschließen", "vorgang", "01.01.2025", "15.01.2025", 15, "Max Mustermann", "Planung", "planned", 0, "1.1"],
        ["1.1.2", "Kickoff-Meeting", "meilenstein", "16.01.2025", "16.01.2025", 1, "Max Mustermann", "Planung", "planned", 0, "1.1"],
    ]
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    col_widths = [12, 35, 15, 13, 13, 13, 20, 18, 15, 14, 22]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Terminplan-Vorlage.xlsx"'},
    )


_TYP_MAP = {
    "vorgang": "vorgang", "task": "vorgang", "aufgabe": "vorgang",
    "meilenstein": "meilenstein", "milestone": "meilenstein",
    "sammelvorgang": "sammelvorgang", "summary": "sammelvorgang",
}

_IMPORT_COL_MAP = {
    "pos_number": ["pos", "pos.", "pos_nr", "pos._nr.", "pos_number", "position", "nummer", "nr", "nr."],
    "title": ["title", "bezeichnung", "beschreibung", "name", "vorgang", "aufgabe", "task"],
    "typ": ["typ", "type", "vorgangstyp"],
    "start_date": ["start", "start_date", "startdatum", "beginn", "von", "anfang"],
    "end_date": ["end", "end_date", "enddatum", "ende", "bis", "fertig"],
    "duration_days": ["duration", "dauer", "duration_days", "dauer_tage", "tage", "dauer_(tage)"],
    "responsible": ["responsible", "verantwortlich", "zuständig", "leiter", "person"],
    "trade": ["trade", "gewerk", "fachbereich", "bereich"],
    "status": ["status"],
    "progress": ["progress", "fortschritt", "fertigstellung", "%", "fortschritt_(%)"],
    "parent_pos_number": ["übergeordnete_pos._nr.", "übergeordnete_pos.__nr.", "parent", "eltern", "übergeordnete_nr"],
}


@router.post("/import-as-version", response_model=schemas.MSPDIImportResult)
async def import_as_version(
    project_id: int = Form(...),
    version_name: str = Form(""),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Projekt nicht gefunden")

    content = await file.read()
    filename = (file.filename or "").lower()
    try:
        if filename.endswith((".xlsx", ".xls")):
            df = pd.read_excel(io.BytesIO(content), dtype=str)
        elif filename.endswith(".csv"):
            for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
                try:
                    df = pd.read_csv(io.BytesIO(content), dtype=str, encoding=enc, sep=None, engine="python")
                    break
                except Exception:
                    continue
            else:
                raise ValueError("CSV konnte nicht gelesen werden")
        else:
            raise ValueError("Nur .xlsx, .xls und .csv werden unterstützt")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    max_ver = db.query(sqlfunc.max(models.ScheduleVersion.version_number)).filter(
        models.ScheduleVersion.project_id == project_id
    ).scalar() or 0
    next_ver = max_ver + 1

    vname = version_name.strip() or f"Import V{next_ver}"
    version = models.ScheduleVersion(
        project_id=project_id,
        name=vname,
        version_number=next_ver,
        is_current=True,
    )
    db.add(version)
    db.flush()

    df.columns = [str(c).strip().lower().replace(" ", "_").replace("-", "_") for c in df.columns]

    def _find_col(target: str) -> Optional[str]:
        for c in _IMPORT_COL_MAP.get(target, [target]):
            if c in df.columns:
                return c
        return None

    errors: list[str] = []
    imported = 0
    skipped = 0
    pos_by_number: dict[str, models.SchedulePosition] = {}
    deferred_parents: list[tuple[models.SchedulePosition, str]] = []

    for idx, row in df.iterrows():
        row_num = int(idx) + 2
        try:
            title_col = _find_col("title")
            if not title_col or pd.isna(row.get(title_col, None)) or str(row.get(title_col, "")).strip() == "":
                skipped += 1
                continue

            def _get(col_name: str) -> Optional[str]:
                c = _find_col(col_name)
                if c is None:
                    return None
                v = row.get(c)
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    return None
                return str(v).strip() or None

            raw_status = _get("status")
            status_val = raw_status.lower().replace(" ", "_") if raw_status else "planned"
            if status_val not in STATUS_VALUES:
                status_val = "planned"

            raw_progress = _get("progress")
            progress_val = 0.0
            if raw_progress:
                try:
                    progress_val = float(raw_progress.replace("%", "").replace(",", "."))
                    if progress_val > 1:
                        progress_val = progress_val / 100
                except ValueError:
                    progress_val = 0.0

            raw_dur = _get("duration_days")
            dur_val = None
            if raw_dur:
                try:
                    dur_val = int(float(raw_dur.replace(",", ".")))
                except ValueError:
                    dur_val = None

            raw_typ = _get("typ")
            typ_val = _TYP_MAP.get((raw_typ or "").lower().strip(), "vorgang")

            pos_number = _get("pos_number")
            parent_pos_number = _get("parent_pos_number")

            pos = models.SchedulePosition(
                version_id=version.id,
                pos_number=pos_number,
                title=str(row[title_col]).strip(),
                start_date=_parse_date(row.get(_find_col("start_date") or "")),
                end_date=_parse_date(row.get(_find_col("end_date") or "")),
                duration_days=dur_val,
                responsible=_get("responsible"),
                trade=_get("trade"),
                typ=typ_val,
                is_milestone=(typ_val == "meilenstein"),
                status=status_val,
                progress=progress_val,
                sort_order=idx,
            )
            db.add(pos)

            if pos_number:
                pos_by_number[pos_number] = pos
            if parent_pos_number:
                deferred_parents.append((pos, parent_pos_number))

            imported += 1
        except Exception as e:
            errors.append(f"Zeile {row_num}: {e}")

    db.flush()

    for pos, parent_num in deferred_parents:
        parent = pos_by_number.get(parent_num)
        if parent:
            pos.parent_id = parent.id
        else:
            errors.append(f"Übergeordnete Pos.-Nr. '{parent_num}' für '{pos.title}' nicht gefunden")

    db.commit()
    return schemas.MSPDIImportResult(
        version_id=version.id,
        positions_created=imported,
        skipped=skipped,
        warnings=errors,
    )


# ── Behinderungsmanagement ────────────────────────────────────────────────────

@router.post("/{position_id}/behinderung/start", response_model=schemas.PositionResponse)
def start_behinderung(position_id: int, db: Session = Depends(get_db)):
    pos = _get_position(db, position_id)
    if pos.behinderung_aktiv:
        raise HTTPException(status_code=400, detail="Behinderung bereits aktiv")
    pos.behinderung_aktiv = True
    pos.behinderung_beginn = datetime.now(timezone.utc).replace(tzinfo=None)
    if pos.status not in ("completed", "cancelled"):
        pos.status = "delayed"
    db.commit()
    db.refresh(pos)
    return pos


@router.post("/{position_id}/behinderung/end", response_model=schemas.PositionResponse)
def end_behinderung(position_id: int, db: Session = Depends(get_db)):
    pos = _get_position(db, position_id)
    if not pos.behinderung_aktiv:
        raise HTTPException(status_code=400, detail="Keine aktive Behinderung")
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    delta_days = max(1, (now - pos.behinderung_beginn).days) if pos.behinderung_beginn else 1
    pos.behinderung_tage_gesamt += delta_days
    pos.behinderung_aktiv = False
    pos.behinderung_beginn = None
    db.commit()
    db.refresh(pos)
    return pos


@router.get("/{position_id}", response_model=schemas.PositionResponse)
def get_position(position_id: int, db: Session = Depends(get_db)):
    return _get_position(db, position_id)
